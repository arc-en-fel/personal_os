import csv
import io
import os
import re
from datetime import datetime, timedelta

import msoffcrypto
from fastapi import FastAPI, Header, HTTPException
from pydantic import BaseModel
from pypdf import PdfReader

app = FastAPI(title="Personal OS statement parser")
MAX_TRANSACTION_AMOUNT = 9_999_999_999.99
DATE_RE = re.compile(r"\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})\b")
AMOUNT_RE = re.compile(r"(?<![\d.])[-(]?(?:\d{1,3}(?:,\d{2,3})+|\d+)(?:\.\d+)?\)?\s*(?:cr|dr)?\b", re.IGNORECASE)

@app.get("/health")
def health(): return {"status": "ok"}

class ParseRequest(BaseModel):
    filename: str
    base64: str
    password: str | None = None

@app.post("/parse")
def parse_statement(payload: ParseRequest, x_parser_token: str | None = Header(default=None)):
    expected = os.environ.get("STATEMENT_PARSER_TOKEN")
    if expected and x_parser_token != expected: raise HTTPException(status_code=401, detail="Invalid parser token")
    try:
        import base64
        raw = base64.b64decode(payload.base64)
        if payload.filename.lower().endswith(".pdf"): rows = rows_from_text(read_pdf(raw, payload.password))
        elif payload.filename.lower().endswith((".xlsx", ".xls")): rows = rows_from_csv(read_excel(raw, payload.password))
        else: rows = rows_from_csv(raw.decode("utf-8-sig"))
        return {"transactions": rows[:500]}
    except Exception as error: raise HTTPException(status_code=422, detail=f"Could not parse statement: {error}") from error

def read_pdf(raw: bytes, password: str | None) -> str:
    reader = PdfReader(io.BytesIO(raw))
    if reader.is_encrypted and (not password or reader.decrypt(password.strip()) == 0): raise ValueError("The file could not be decrypted with this password")
    return "\n".join(page.extract_text(extraction_mode="layout") or "" for page in reader.pages)

def read_excel(raw: bytes, password: str | None) -> str:
    source = io.BytesIO(raw)
    if password:
        decrypted = io.BytesIO(); office = msoffcrypto.OfficeFile(source); office.load_key(password=password.strip()); office.decrypt(decrypted); source = decrypted
    from openpyxl import load_workbook
    workbook = load_workbook(source, read_only=True, data_only=True); sheets = [list(sheet.iter_rows(values_only=True)) for sheet in workbook.worksheets]
    rows = max(sheets, key=lambda sheet: sum(any(cell not in (None, "") for cell in row) for row in sheet), default=[])
    output = io.StringIO(); csv.writer(output).writerows(rows); return output.getvalue()

def rows_from_csv(value: str) -> list[dict]:
    rows = [row for row in csv.reader(io.StringIO(value)) if any(cell.strip() for cell in row)]
    if not rows: return []
    header_index = next((index for index, row in enumerate(rows[:20]) if is_header(row)), None)
    if header_index is not None:
        headers = [cell.strip().lower() for cell in rows[header_index]]
        date_index = find_column(headers, ("transaction date", "txn date", "value date", "date")); debit_index = find_column(headers, ("debit", "withdrawal", "outflow")); credit_index = find_column(headers, ("credit", "deposit", "inflow")); amount_index = find_column(headers, ("amount", "transaction amount")); merchant_index = find_column(headers, ("merchant", "description", "narration", "particular", "details"))
        return mapped_rows(rows[header_index + 1:], date_index, debit_index, credit_index, amount_index, merchant_index)
    return inferred_rows(rows)

def is_header(row: list[str]) -> bool:
    cells = [cell.strip().lower() for cell in row]; return any("date" in cell or "txn" in cell for cell in cells) and any(any(word in cell for word in ("amount", "debit", "credit", "withdraw", "deposit", "balance")) for cell in cells)

def mapped_rows(rows: list[list[str]], date_index: int | None, debit_index: int | None, credit_index: int | None, amount_index: int | None, merchant_index: int | None) -> list[dict]:
    result = []
    for row in rows:
        date = normalize_date(row[date_index]) if date_index is not None and date_index < len(row) else first_date(row)
        if not date: continue
        if debit_index is not None or credit_index is not None:
            debit = parse_amount(row[debit_index]) if debit_index is not None and debit_index < len(row) else 0; credit = parse_amount(row[credit_index]) if credit_index is not None and credit_index < len(row) else 0; amount = -abs(debit) if debit else abs(credit)
        else: amount = parse_amount(row[amount_index]) if amount_index is not None and amount_index < len(row) else first_amount(row)
        merchant = row[merchant_index].strip() if merchant_index is not None and merchant_index < len(row) else first_text(row, date)
        if amount and abs(amount) <= MAX_TRANSACTION_AMOUNT: result.append(transaction(date, amount, merchant))
    return result

def inferred_rows(rows: list[list[str]]) -> list[dict]:
    result = []
    for row in rows:
        date = first_date(row); amount = first_amount(row)
        if date and amount: result.append(transaction(date, amount, first_text(row, date)))
    return result

def first_date(row: list[str]) -> str | None:
    for cell in row:
        date = normalize_date(cell)
        if date: return date
    return None

def first_amount(row: list[str]) -> float:
    for cell in row:
        amount = parse_amount(cell)
        if amount and abs(amount) <= MAX_TRANSACTION_AMOUNT: return amount
    return 0

def first_text(row: list[str], date: str | None) -> str:
    for cell in row:
        text = cell.strip()
        if text and normalize_date(text) != date and not parse_amount(text): return text
    return ""

def find_column(headers: list[str], names: tuple[str, ...]) -> int | None:
    for index, header in enumerate(headers):
        if any(name in header for name in names): return index
    return None

def rows_from_text(value: str) -> list[dict]:
    lines = [" ".join(line.split()) for line in value.splitlines() if line.strip()]
    result = []; index = 0
    while index < len(lines):
        date_match = DATE_RE.search(lines[index])
        if not date_match: index += 1; continue
        date_text = date_match.group(0); date = normalize_date(date_text); window = lines[index:index + 3]; joined = " ".join(window)
        without_date = joined.replace(date_text, " ", 1); amounts = list(AMOUNT_RE.finditer(without_date)); amount_match = amounts[0] if amounts else None
        if amount_match and date:
            amount = parse_amount(amount_match.group(0))
            if amount and abs(amount) <= MAX_TRANSACTION_AMOUNT:
                merchant = without_date[:amount_match.start()].strip(" -|:")
                if merchant and not any(word in merchant.lower() for word in ("date", "amount", "balance", "debit", "credit")): result.append(transaction(date, amount, merchant))
        index += 1
    return result

def transaction(date: str, amount: float, merchant: str | None) -> dict: return {"amount": abs(amount), "transaction_type": "expense" if amount < 0 else "income", "merchant": merchant or None, "transaction_date": date}

def parse_amount(value: str) -> float:
    text = str(value).strip().lower()
    if not text: return 0
    negative = "-" in text or text.endswith("dr") or (text.startswith("(") and text.endswith(")")); clean = re.sub(r"[^0-9.,]", "", text).replace(",", "")
    if clean.count(".") > 1:
        last_separator = clean.rfind("."); decimal_part = clean[last_separator + 1:]; clean = clean[:last_separator].replace(".", "") + ("." + decimal_part if len(decimal_part) <= 2 else "")
    try: number = float(clean or 0)
    except ValueError: return 0
    number = -abs(number) if negative else abs(number); return number if abs(number) <= MAX_TRANSACTION_AMOUNT else 0

def normalize_date(value: str) -> str | None:
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y"):
        try: return datetime.strptime(text, pattern).date().isoformat()
        except ValueError: continue
    try:
        serial = float(text)
        if 20_000 <= serial <= 80_000: return (datetime(1899, 12, 30) + timedelta(days=serial)).date().isoformat()
    except ValueError: pass
    return None
